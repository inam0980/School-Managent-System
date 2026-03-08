from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.http import HttpResponse
from datetime import date, timedelta
from .models import Invoice, InvoiceItem, Payment, FeeCategory, Discount
from students.models import Student
from .utils import generate_invoice_pdf
from accounts.decorators import module_required, action_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from io import BytesIO
import qrcode
import base64

# Create your views here.

@module_required('billing')
def invoice_list(request):
    """
    Display list of all invoices with search and filters
    Protected: Requires 'billing' module access
    """
    invoices = Invoice.objects.select_related('student').all()
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        invoices = invoices.filter(
            Q(invoice_number__icontains=search_query) |
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query) |
            Q(student__first_name_arabic__icontains=search_query) |
            Q(student__last_name_arabic__icontains=search_query) |
            Q(student__student_id__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        invoices = invoices.filter(status=status_filter)
    
    # Filter by date range
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        invoices = invoices.filter(invoice_date__gte=date_from)
    if date_to:
        invoices = invoices.filter(invoice_date__lte=date_to)
    
    # Handle export requests
    export_format = request.GET.get('export', '')
    if export_format == 'excel':
        return export_invoices_excel(invoices)
    elif export_format == 'pdf':
        return export_invoices_pdf(invoices)
    
    # Calculate statistics
    stats = Invoice.objects.aggregate(
        total_revenue=Sum('total_amount'),
        total_paid=Sum('paid_amount'),
        total_pending=Sum('balance_amount'),
        total_count=Count('id')
    )
    
    # Count by status
    pending_count = Invoice.objects.filter(status='pending').count()
    overdue_count = Invoice.objects.filter(status='overdue').count()
    paid_count = Invoice.objects.filter(status='paid').count()
    
    context = {
        'invoices': invoices,
        'search_query': search_query,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'total_revenue': stats['total_revenue'] or 0,
        'total_paid': stats['total_paid'] or 0,
        'total_pending': stats['total_pending'] or 0,
        'total_count': stats['total_count'] or 0,
        'pending_count': pending_count,
        'overdue_count': overdue_count,
        'paid_count': paid_count,
    }
    
    return render(request, 'billing/invoice_list.html', context)


@login_required
def invoice_detail(request, invoice_number):
    """Display detailed invoice information"""
    invoice = get_object_or_404(Invoice, invoice_number=invoice_number)
    items = invoice.items.select_related('fee_category').all()
    payments = invoice.payments.all()
    
    context = {
        'invoice': invoice,
        'items': items,
        'payments': payments,
    }
    
    return render(request, 'billing/invoice_detail.html', context)


@login_required
@module_required('billing')
def invoice_create(request):
    """Create new invoice"""
    if request.method == 'POST':
        try:
            # Get student
            student_id = request.POST.get('student_id')
            student = Student.objects.get(student_id=student_id)
            
            # Parse dates
            from datetime import datetime
            invoice_date_str = request.POST.get('invoice_date')
            due_date_str = request.POST.get('due_date')
            
            invoice_date = datetime.strptime(invoice_date_str, '%Y-%m-%d').date() if invoice_date_str else date.today()
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date() if due_date_str else date.today() + timedelta(days=30)
            
            # Create invoice
            invoice = Invoice.objects.create(
                student=student,
                academic_year=request.POST.get('academic_year', '2024-2025'),
                invoice_date=invoice_date,
                due_date=due_date,
                notes=request.POST.get('notes', ''),
                created_by=request.user
            )
            
            # Get discount if selected
            discount_id = request.POST.get('discount')
            if discount_id:
                invoice.discount = Discount.objects.get(id=discount_id)
                invoice.save()
            
            # Create invoice items
            fee_categories = request.POST.getlist('fee_category[]')
            descriptions = request.POST.getlist('description[]')
            quantities = request.POST.getlist('quantity[]')
            unit_prices = request.POST.getlist('unit_price[]')
            
            for i in range(len(fee_categories)):
                if fee_categories[i]:
                    InvoiceItem.objects.create(
                        invoice=invoice,
                        fee_category_id=fee_categories[i],
                        description=descriptions[i] if i < len(descriptions) else '',
                        quantity=int(quantities[i]) if i < len(quantities) else 1,
                        unit_price=float(unit_prices[i]) if i < len(unit_prices) else 0
                    )
            
            # Calculate totals
            invoice.calculate_totals()
            
            messages.success(request, f'Invoice {invoice.invoice_number} created successfully!')
            return redirect('billing:detail', invoice_number=invoice.invoice_number)
            
        except Exception as e:
            messages.error(request, f'Error creating invoice: {str(e)}')
            return render(request, 'billing/invoice_create.html', get_create_context())
    
    context = get_create_context()
    return render(request, 'billing/invoice_create.html', context)


def get_create_context():
    """Helper function to get context for invoice creation"""
    students = Student.objects.filter(is_active=True).order_by('first_name', 'last_name')
    fee_categories = FeeCategory.objects.filter(is_active=True).order_by('category_name')
    discounts = Discount.objects.filter(is_active=True).order_by('discount_name')
    
    return {
        'students': students,
        'fee_categories': fee_categories,
        'discounts': discounts,
        'today': date.today(),
        'default_due_date': date.today() + timedelta(days=30),
    }


@login_required
@module_required('billing')
def simplified_invoice_create(request):
    """Create simplified tax invoice (no VAT)"""
    if request.method == 'POST':
        # TODO: Implement simplified invoice creation logic
        messages.info(request, 'Simplified invoice creation coming soon!')
        return redirect('billing:list')
    
    context = get_create_context()
    return render(request, 'billing/simplified_invoice_create.html', context)


@login_required
@module_required('billing')
def vat_invoice_create(request):
    """Create VAT tax invoice (15% VAT)"""
    if request.method == 'POST':
        # TODO: Implement VAT invoice creation logic
        messages.info(request, 'VAT invoice creation coming soon!')
        return redirect('billing:list')
    
    context = get_create_context()
    return render(request, 'billing/vat_invoice_create.html', context)


@login_required
@login_required
def payment_create(request, invoice_number):
    """Record payment for an invoice"""
    invoice = get_object_or_404(Invoice, invoice_number=invoice_number)
    
    if request.method == 'POST':
        try:
            amount = float(request.POST.get('amount'))
            
            # Validate amount
            if amount <= 0:
                raise ValueError("Payment amount must be greater than zero")
            if amount > invoice.balance_amount:
                raise ValueError("Payment amount cannot exceed balance amount")
            
            # Create payment
            payment = Payment.objects.create(
                invoice=invoice,
                payment_date=request.POST.get('payment_date'),
                amount=amount,
                payment_method=request.POST.get('payment_method'),
                reference_number=request.POST.get('reference_number', ''),
                cheque_number=request.POST.get('cheque_number', ''),
                bank_name=request.POST.get('bank_name', ''),
                notes=request.POST.get('notes', ''),
                received_by=request.user
            )
            
            messages.success(request, f'Payment {payment.payment_number} recorded successfully! Receipt: {payment.receipt_number}')
            return redirect('billing:detail', invoice_number=invoice_number)
            
        except Exception as e:
            messages.error(request, f'Error recording payment: {str(e)}')
            return redirect('billing:payment', invoice_number=invoice_number)
    
    context = {
        'invoice': invoice,
        'today': date.today(),
    }
    
    return render(request, 'billing/payment_create.html', context)


@login_required
@module_required('billing')
def payment_list(request):
    """Display list of all payments"""
    payments = Payment.objects.select_related('invoice', 'invoice__student').all()
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        payments = payments.filter(
            Q(payment_number__icontains=search_query) |
            Q(receipt_number__icontains=search_query) |
            Q(invoice__invoice_number__icontains=search_query) |
            Q(invoice__student__first_name__icontains=search_query) |
            Q(invoice__student__last_name__icontains=search_query)
        )
    
    # Filter by payment method
    method_filter = request.GET.get('method', '')
    if method_filter:
        payments = payments.filter(payment_method=method_filter)
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        payments = payments.filter(status=status_filter)
    
    # Calculate statistics
    stats = Payment.objects.filter(status='completed').aggregate(
        total_collected=Sum('amount'),
        total_count=Count('id')
    )
    
    # Calculate average payment
    total_collected = stats['total_collected'] or 0
    total_count = stats['total_count'] or 0
    average_payment = total_collected / total_count if total_count > 0 else 0
    
    context = {
        'payments': payments,
        'search_query': search_query,
        'method_filter': method_filter,
        'status_filter': status_filter,
        'total_collected': total_collected,
        'total_count': total_count,
        'average_payment': average_payment,
    }
    
    return render(request, 'billing/payment_list.html', context)


@login_required
def invoice_pdf(request, invoice_number):
    """Generate and download invoice PDF"""
    invoice = get_object_or_404(Invoice, invoice_number=invoice_number)
    
    try:
        # Generate PDF
        pdf_buffer = generate_invoice_pdf(invoice)
        
        # Create HTTP response
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Invoice_{invoice_number}.pdf"'
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error generating PDF: {str(e)}')
        return redirect('billing:detail', invoice_number=invoice_number)


@login_required
def invoice_print(request, invoice_number):
    """Display printable invoice view"""
    invoice = get_object_or_404(Invoice, invoice_number=invoice_number)
    items = invoice.items.select_related('fee_category').all()
    payments = invoice.payments.all()
    
    # Generate QR Code with invoice information
    qr_data = f"""Invoice: {invoice.invoice_number}
Student: {invoice.student.get_full_name()}
Date: {invoice.invoice_date.strftime('%d/%m/%Y')}
Amount: {invoice.total_amount} SAR
VAT: {invoice.vat_amount} SAR
Status: {invoice.get_status_display()}"""
    
    # Create QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)
    
    # Generate QR code image
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Convert to base64 for embedding in HTML
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()
    qr_code_data_uri = f"data:image/png;base64,{qr_code_base64}"
    
    context = {
        'invoice': invoice,
        'items': items,
        'payments': payments,
        'qr_code': qr_code_data_uri,
    }
    
    return render(request, 'billing/invoice_print.html', context)


# Export Functions
def export_invoices_excel(invoices):
    """Export invoices to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    
    # Define styles
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Invoice Number', 'Student ID', 'Student Name', 'Academic Year',
        'Invoice Date', 'Due Date', 'Subtotal', 'Discount', 'VAT', 
        'Total Amount', 'Paid Amount', 'Balance', 'Status'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Write data
    for row, invoice in enumerate(invoices, 2):
        data = [
            invoice.invoice_number,
            invoice.student.student_id,
            invoice.student.get_full_name(),
            invoice.academic_year,
            invoice.invoice_date.strftime('%d/%m/%Y'),
            invoice.due_date.strftime('%d/%m/%Y'),
            float(invoice.subtotal),
            float(invoice.discount_amount),
            float(invoice.vat_amount),
            float(invoice.total_amount),
            float(invoice.paid_amount),
            float(invoice.balance_amount),
            invoice.status.upper()
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left' if col <= 4 else 'right', vertical='center')
            
            # Color code status
            if col == 13:  # Status column
                if value == 'PAID':
                    cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    cell.font = Font(color="065F46", bold=True)
                elif value == 'OVERDUE':
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    cell.font = Font(color="991B1B", bold=True)
                elif value == 'PENDING':
                    cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                    cell.font = Font(color="1E40AF", bold=True)
    
    # Adjust column widths
    column_widths = [18, 12, 25, 12, 12, 12, 12, 12, 10, 12, 12, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Invoices_{date.today().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response


def export_invoices_pdf(invoices):
    """Export invoices to PDF file"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                           leftMargin=0.5*inch, rightMargin=0.5*inch,
                           topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph("<b>Invoices Report</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Prepare data
    data = [[
        'Invoice #', 'Student ID', 'Student Name', 'Date', 'Due Date',
        'Total', 'Paid', 'Balance', 'Status'
    ]]
    
    for invoice in invoices:
        data.append([
            invoice.invoice_number,
            invoice.student.student_id,
            invoice.student.get_full_name()[:20],  # Truncate long names
            invoice.invoice_date.strftime('%d/%m/%Y'),
            invoice.due_date.strftime('%d/%m/%Y'),
            f"{invoice.total_amount:.2f}",
            f"{invoice.paid_amount:.2f}",
            f"{invoice.balance_amount:.2f}",
            invoice.status.upper()
        ])
    
    # Create table
    table = Table(data, colWidths=[1.2*inch, 1*inch, 1.8*inch, 0.9*inch, 0.9*inch, 
                                   0.9*inch, 0.9*inch, 0.9*inch, 0.8*inch])
    
    # Style table
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data rows
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (2, -1), 'LEFT'),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Create response
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Invoices_{date.today().strftime("%Y%m%d")}.pdf"'
    
    return response

